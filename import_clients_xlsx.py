r"""Import a Smart Scheduling Clients.xlsx export into A2Z.

Run: .\.venv\Scripts\python.exe import_clients_xlsx.py clients.xlsx
The source ID is stable, so rerunning updates instead of duplicating clients.
Imported clients never receive login access, and email is optional.
"""

from __future__ import annotations

import re
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from database import database_path, init_db


HEADERS = (
    "Id", "First Name", "Last Name", "Email", "Secondary Email", "Phone",
    "Secondary Phone", "Birthday", "Gender", "Street", "City", "Zip Code",
    "Created", "Notes",
)


def clean(value) -> str:
    return " ".join(str(value).split()).strip() if value is not None else ""


def birthday_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = clean(value)
    for pattern in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, pattern).date().isoformat()
        except ValueError:
            pass
    return None


def created_value(value):
    return value.isoformat(sep=" ", timespec="seconds") if isinstance(value, datetime) else None


def import_clients(source, *, branch_id=None, source_namespace="default", preserve_existing=False):
    source = Path(source).expanduser().resolve()
    if not source.is_file():
        raise ValueError(f"Workbook not found: {source}")

    init_db()
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["Clients"] if "Clients" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    missing = [name for name in HEADERS if name not in headers]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")
    column = {name: headers.index(name) for name in HEADERS}

    connection = sqlite3.connect(database_path())
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    branch = connection.execute(
        "SELECT id FROM branches WHERE id=? AND is_active=1" if branch_id is not None
        else "SELECT id FROM branches WHERE is_active=1 ORDER BY id LIMIT 1",
        (branch_id,) if branch_id is not None else (),
    ).fetchone()
    if not branch:
        raise ValueError("No active branch exists.")

    added = updated = 0
    try:
        connection.execute("BEGIN IMMEDIATE")
        for row_number, row in enumerate(rows, 2):
            source_id = clean(row[column["Id"]]) or f"row-{row_number}"
            stable_id = re.sub(r"[^0-9A-Za-z_-]+", "-", source_id).strip("-") or str(row_number)
            namespace = re.sub(r"[^0-9A-Za-z_-]+", "-", source_namespace).strip("-").lower() or "default"
            username = f"smart-client-{namespace}-{stable_id}".lower()
            first = clean(row[column["First Name"]])
            last = clean(row[column["Last Name"]])
            full_name = " ".join(value for value in (first, last) if value) or f"Client {source_id}"
            email = clean(row[column["Email"]]) or None
            phone = clean(row[column["Phone"]]) or None
            secondary_email = clean(row[column["Secondary Email"]]) or None
            secondary_phone = clean(row[column["Secondary Phone"]]) or None
            gender = clean(row[column["Gender"]]).lower()
            gender = gender if gender in {"male", "female", "other"} else None
            existing = connection.execute(
                "SELECT id FROM users WHERE lower(username)=lower(?)", (username,)
            ).fetchone()
            if existing:
                user_id = existing["id"]
                if not preserve_existing:
                    connection.execute(
                        "UPDATE users SET full_name=?,email=?,phone=?,branch_id=?,is_active=1,login_enabled=0,must_change_password=0,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                        (full_name, email, phone, branch["id"], user_id),
                    )
                updated += 1
            else:
                user_id = connection.execute(
                    """INSERT INTO users
                       (username,password_hash,role,full_name,email,phone,branch_id,
                        is_active,login_enabled,must_change_password,created_at,updated_at)
                       VALUES (?,'!','student',?,?,?,?,1,0,0,COALESCE(?,CURRENT_TIMESTAMP),CURRENT_TIMESTAMP)""",
                    (username, full_name, email, phone, branch["id"], created_value(row[column["Created"]])),
                ).lastrowid
                added += 1
            if existing and preserve_existing:
                continue
            connection.execute(
                """INSERT INTO client_profiles
                   (user_id,secondary_phone,secondary_email,birthday,gender,zip_code,
                    city,street,internal_notes,reminders_enabled,preferred_channel,
                    source_reference,updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,1,?,?,CURRENT_TIMESTAMP)
                   ON CONFLICT(user_id) DO UPDATE SET
                    secondary_phone=excluded.secondary_phone,
                    secondary_email=excluded.secondary_email,
                    birthday=excluded.birthday,gender=excluded.gender,
                    zip_code=excluded.zip_code,city=excluded.city,street=excluded.street,
                    internal_notes=excluded.internal_notes,
                    preferred_channel=excluded.preferred_channel,
                    source_reference=excluded.source_reference,updated_at=CURRENT_TIMESTAMP""",
                (
                    user_id, secondary_phone, secondary_email,
                    birthday_value(row[column["Birthday"]]), gender,
                    clean(row[column["Zip Code"]]) or None,
                    clean(row[column["City"]]) or None,
                    clean(row[column["Street"]]) or None,
                    clean(row[column["Notes"]]) or None,
                    "email" if email else "sms", f"smart:{namespace}:client:{source_id}",
                ),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
        workbook.close()
    return added, updated


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: import_clients_xlsx.py CLIENTS.xlsx")
    try:
        added, updated = import_clients(sys.argv[1])
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc
    print(f"Client import complete: {added} added, {updated} updated.")


if __name__ == "__main__":
    main()
