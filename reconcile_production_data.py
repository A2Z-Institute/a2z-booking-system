"""Safely import the latest cleaned Smart Scheduling export into production.

The operation is deliberately manual and idempotent. It creates and verifies a
database backup before importing, preserves Smart Scheduling source IDs, merges
only strong duplicate-client matches, and cancels (rather than deletes) exact
duplicate upcoming appointments so booking history remains recoverable.

Production usage inside the Coolify application container::

    python reconcile_production_data.py --apply
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from backup_database import create_backup, verify_database
from database import database_path, get_db, init_db
from import_smartscheduling_backup import import_backup


DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parent
    / "data-import"
    / "A2Z_Clean_Upcoming_Import_2026-08-19.xlsx"
)
ACTIVE_STATUSES = {
    "Pending",
    "Approved",
    "Not Confirmed",
    "Running Late",
    "Arrived",
    "Rescheduled",
    "No Action",
}
PROFILE_FIELDS = (
    "secondary_phone",
    "secondary_email",
    "birthday",
    "gender",
    "zip_code",
    "city",
    "street",
    "tags",
)


def clean(value) -> str:
    return " ".join(str(value).split()).strip() if value is not None else ""


def normalise_text(value) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", clean(value).upper()).strip()


def normalise_phone(value) -> str:
    return "".join(re.findall(r"\d", clean(value)))


def client_key(full_name, phone, email):
    name = normalise_text(full_name)
    phone_key = normalise_phone(phone)
    email_key = clean(email).casefold()
    if name and phone_key:
        return ("name+phone", name, phone_key)
    if name and email_key:
        return ("name+email", name, email_key)
    return None


def canonical_source_ids(workbook_path: Path) -> set[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Clients"]
        rows = sheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]
        id_index = headers.index("Id")
        return {
            clean(row[id_index])
            for row in rows
            if id_index < len(row) and clean(row[id_index])
        }
    finally:
        workbook.close()


def preflight_instructors(workbook_path: Path) -> dict:
    """Confirm every scheduling sheet maps to exactly one live instructor."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet_names = [
            name for name in workbook.sheetnames if name not in {"Clients", "Services"}
        ]
    finally:
        workbook.close()

    with get_db() as conn:
        branch = conn.execute(
            "SELECT id FROM branches WHERE is_active = 1 ORDER BY id LIMIT 1"
        ).fetchone()
        if not branch:
            return {"matched": [], "missing": sheet_names, "ambiguous": []}
        live = [
            (int(row["id"]), normalise_text(row["name"]))
            for row in conn.execute(
                "SELECT id, name FROM instructors WHERE branch_id = ? AND is_active = 1",
                (branch["id"],),
            )
        ]

    matched, missing, ambiguous = [], [], []
    for sheet_name in sheet_names:
        sheet_key = normalise_text(sheet_name.replace("...", ""))
        exact = [item for item in live if item[1] == sheet_key]
        candidates = exact or [
            item
            for item in live
            if item[1].startswith(sheet_key) or sheet_key.startswith(item[1][:20])
        ]
        unique = {item[0]: item for item in candidates}
        if len(unique) == 1:
            instructor_id, instructor_name = next(iter(unique.values()))
            matched.append(
                {
                    "sheet": sheet_name,
                    "instructor_id": instructor_id,
                    "instructor_name": instructor_name,
                }
            )
        elif not unique:
            missing.append(sheet_name)
        else:
            ambiguous.append(
                {
                    "sheet": sheet_name,
                    "candidate_ids": sorted(unique),
                }
            )
    return {"matched": matched, "missing": missing, "ambiguous": ambiguous}


def _merge_profile(conn, canonical_id: int, duplicate_id: int) -> list[str]:
    canonical = conn.execute(
        "SELECT * FROM client_profiles WHERE user_id = ?", (canonical_id,)
    ).fetchone()
    duplicate = conn.execute(
        "SELECT * FROM client_profiles WHERE user_id = ?", (duplicate_id,)
    ).fetchone()
    if not duplicate:
        return []
    if not canonical:
        conn.execute(
            "UPDATE client_profiles SET user_id = ? WHERE user_id = ?",
            (canonical_id, duplicate_id),
        )
        return ["client_profile"]

    updates = {}
    for field in PROFILE_FIELDS:
        if not clean(canonical[field]) and clean(duplicate[field]):
            updates[field] = duplicate[field]
    canonical_notes = clean(canonical["internal_notes"])
    duplicate_notes = clean(duplicate["internal_notes"])
    if duplicate_notes and duplicate_notes not in canonical_notes:
        updates["internal_notes"] = (
            f"{canonical_notes}\n{duplicate_notes}".strip()
        )
    if updates:
        assignments = ", ".join(f"{field} = ?" for field in updates)
        conn.execute(
            f"UPDATE client_profiles SET {assignments}, "
            "updated_at = CURRENT_TIMESTAMP WHERE user_id = ?",
            (*updates.values(), canonical_id),
        )
    return sorted(updates)


def merge_duplicate_clients(
    conn: sqlite3.Connection, retained_source_ids: set[str]
) -> list[dict]:
    rows = conn.execute(
        """
        SELECT u.id, u.username, u.full_name, u.email, u.phone, u.is_active,
               cp.source_reference,
               (SELECT count(*) FROM bookings b WHERE b.student_user_id = u.id)
                   AS booking_count
        FROM users u
        LEFT JOIN client_profiles cp ON cp.user_id = u.id
        WHERE u.role = 'student'
        ORDER BY u.id
        """
    ).fetchall()
    groups = defaultdict(list)
    for row in rows:
        match = client_key(row["full_name"], row["phone"], row["email"])
        if match:
            groups[match].append(row)

    merged = []
    for match, members in groups.items():
        if len(members) < 2:
            continue

        def score(row):
            retained_source = clean(row["source_reference"]) in retained_source_ids
            return (
                int(retained_source),
                int(row["is_active"]),
                int(row["booking_count"] or 0),
                -int(row["id"]),
            )

        canonical = max(members, key=score)
        for duplicate in members:
            if duplicate["id"] == canonical["id"]:
                continue
            profile_fields = _merge_profile(conn, canonical["id"], duplicate["id"])

            assignments = conn.execute(
                """
                SELECT instructor_id, assigned_by, is_active, assigned_at, ended_at
                FROM student_instructor_assignments
                WHERE student_user_id = ?
                """,
                (duplicate["id"],),
            ).fetchall()
            for assignment in assignments:
                conn.execute(
                    """
                    INSERT INTO student_instructor_assignments
                        (student_user_id, instructor_id, assigned_by, is_active,
                         assigned_at, ended_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(student_user_id, instructor_id) DO UPDATE SET
                        is_active = max(student_instructor_assignments.is_active,
                                        excluded.is_active),
                        ended_at = CASE
                            WHEN excluded.is_active = 1 THEN NULL
                            ELSE student_instructor_assignments.ended_at
                        END
                    """,
                    (
                        canonical["id"],
                        assignment["instructor_id"],
                        assignment["assigned_by"],
                        assignment["is_active"],
                        assignment["assigned_at"],
                        assignment["ended_at"],
                    ),
                )
            conn.execute(
                "DELETE FROM student_instructor_assignments WHERE student_user_id = ?",
                (duplicate["id"],),
            )
            moved_bookings = conn.execute(
                "UPDATE bookings SET student_user_id = ? WHERE student_user_id = ?",
                (canonical["id"], duplicate["id"]),
            ).rowcount
            conn.execute(
                """
                UPDATE users
                SET is_active = 0, login_enabled = 0,
                    deactivated_at = COALESCE(deactivated_at, CURRENT_TIMESTAMP),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (duplicate["id"],),
            )
            conn.execute(
                """
                INSERT INTO audit_events(event_type, details)
                VALUES ('duplicate_client_merged', ?)
                """,
                (
                    json.dumps(
                        {
                            "canonical_user_id": canonical["id"],
                            "duplicate_user_id": duplicate["id"],
                            "match_rule": match[0],
                            "bookings_reassigned": moved_bookings,
                        },
                        separators=(",", ":"),
                    ),
                ),
            )
            merged.append(
                {
                    "canonical_user_id": canonical["id"],
                    "canonical_source_reference": canonical["source_reference"],
                    "duplicate_user_id": duplicate["id"],
                    "duplicate_source_reference": duplicate["source_reference"],
                    "match_rule": match[0],
                    "bookings_reassigned": moved_bookings,
                    "profile_fields_merged": profile_fields,
                }
            )
    return merged


def booking_key(row):
    return (
        int(row["branch_id"]),
        int(row["instructor_id"]),
        int(row["machine_id"]),
        row["target_date"],
        row["start_time"],
        row["end_time"],
        normalise_text(row["student_name"]),
        normalise_phone(row["mobile_number"]),
        normalise_text(row["service_name"]),
    )


def cancel_exact_duplicate_upcoming_bookings(
    conn: sqlite3.Connection, cutoff: str
) -> list[dict]:
    placeholders = ",".join("?" for _ in ACTIVE_STATUSES)
    rows = conn.execute(
        f"""
        SELECT * FROM bookings
        WHERE target_date >= ? AND validation_status IN ({placeholders})
        ORDER BY id
        """,
        (cutoff, *sorted(ACTIVE_STATUSES)),
    ).fetchall()
    groups = defaultdict(list)
    for row in rows:
        groups[booking_key(row)].append(row)

    cancelled = []
    for members in groups.values():
        if len(members) < 2:
            continue
        canonical = max(
            members,
            key=lambda row: (
                int(bool(clean(row["source_reference"]))),
                int(bool(row["student_user_id"])),
                int(bool(clean(row["mobile_number"]))),
                -int(row["id"]),
            ),
        )
        for duplicate in members:
            if duplicate["id"] == canonical["id"]:
                continue
            marker = f"Reconciliation: duplicate of booking #{canonical['id']}"
            notes = clean(duplicate["notes"])
            if marker not in notes:
                notes = f"{notes}\n{marker}".strip()
            conn.execute(
                """
                UPDATE bookings
                SET validation_status = 'Cancelled', cancelled_at = CURRENT_TIMESTAMP,
                    notes = ?, calendar_revision = calendar_revision + 1,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (notes, duplicate["id"]),
            )
            conn.execute(
                """
                INSERT INTO audit_events(booking_id, event_type, details)
                VALUES (?, 'duplicate_booking_cancelled_by_reconciliation', ?)
                """,
                (
                    duplicate["id"],
                    json.dumps(
                        {"canonical_booking_id": canonical["id"]},
                        separators=(",", ":"),
                    ),
                ),
            )
            cancelled.append(
                {
                    "canonical_booking_id": canonical["id"],
                    "cancelled_duplicate_booking_id": duplicate["id"],
                    "date": duplicate["target_date"],
                    "start": duplicate["start_time"],
                    "end": duplicate["end_time"],
                    "client": duplicate["student_name"],
                    "instructor_id": duplicate["instructor_id"],
                    "machine_id": duplicate["machine_id"],
                }
            )
    return cancelled


def write_report(report_path: Path, report: dict) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Apply the import")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--database")
    parser.add_argument("--report")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.is_file():
        raise SystemExit(f"Clean import workbook not found: {workbook_path}")
    if args.database:
        os.environ["A2Z_DATABASE"] = str(Path(args.database).expanduser().resolve())
    retained_ids = canonical_source_ids(workbook_path)
    cutoff = "2026-08-19"
    report_path = Path(
        args.report
        or (
            Path(os.environ.get("A2Z_BACKUP_DIR", "/data/backups"))
            / f"reconciliation-{datetime.now():%Y%m%d-%H%M%S}.json"
        )
    ).resolve()

    report = {
        "applied": bool(args.apply),
        "database": str(database_path()),
        "workbook": str(workbook_path),
        "cutoff": cutoff,
        "retained_client_source_ids": len(retained_ids),
        "started_at": datetime.now(timezone.utc).isoformat(),
    }
    database_file = database_path()
    if not database_file.is_file():
        raise SystemExit(f"Live database not found: {database_file}")
    verify_database(database_file)
    instructor_preflight = preflight_instructors(workbook_path)
    report["instructor_preflight"] = instructor_preflight
    if instructor_preflight["missing"] or instructor_preflight["ambiguous"]:
        report["message"] = (
            "Stopped before import: one or more workbook sheets do not map "
            "to exactly one active instructor."
        )
        write_report(report_path, report)
        print(json.dumps(report, indent=2))
        return 2
    if not args.apply:
        report["message"] = (
            "Dry run only. Run again with --apply after confirming the database backup."
        )
        write_report(report_path, report)
        print(json.dumps(report, indent=2))
        return 0

    backup_path = create_backup(once_per_day=False, retain=int(os.environ.get("A2Z_BACKUP_RETENTION", "30")))
    if not backup_path:
        raise SystemExit("A verified pre-import backup could not be created.")
    report["pre_import_backup"] = str(backup_path)
    init_db()
    verify_database(database_path())

    import_counts, import_errors = import_backup(workbook_path, database_path())
    report["import_counts"] = import_counts
    report["import_errors"] = import_errors
    if import_errors:
        report["completed_at"] = datetime.now(timezone.utc).isoformat()
        report["message"] = "Import completed with errors; duplicate reconciliation was not applied."
        write_report(report_path, report)
        print(json.dumps(report, indent=2))
        return 2

    with get_db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        merged_clients = merge_duplicate_clients(conn, retained_ids)
        cancelled_bookings = cancel_exact_duplicate_upcoming_bookings(conn, cutoff)
    verify_database(database_path())

    report["duplicate_clients_merged"] = merged_clients
    report["duplicate_upcoming_bookings_cancelled"] = cancelled_bookings
    report["completed_at"] = datetime.now(timezone.utc).isoformat()
    report["integrity_check"] = "ok"
    write_report(report_path, report)
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
