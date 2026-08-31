"""Restore a Smart Scheduling XLSX backup into an A2Z database.

The import is idempotent: each exported row keeps a stable source reference.
Appointments, long booking-slot bands, and breaks/busy time are restored as
different record types so appointments may remain inside their booking slots.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from database import database_path, init_db
from import_clients_xlsx import import_clients


BREAK_NAMES = {"BREAK", "BREAKFAST", "TEA BREAK", "LUNCH BREAK", "NO BOOKING", "LEAVE"}
STATUS_MAP = {
    "": "Approved", "NONE": "Approved", "CONFIRMED": "Approved",
    "NOT CONFIRMED": "Pending", "PENDING": "Pending",
    "COMPLETED": "Completed", "CANCELLED": "Cancelled",
    "NO-SHOW": "No-show", "NO SHOW": "No-show",
    "RESCHEDULED": "Rescheduled", "RUNNING LATE": "Running Late",
    "ARRIVED": "Arrived",
}
COLORS = ("#E6C8F5", "#7CC5E3", "#F5E39B", "#B9E5BF", "#F1DCC0", "#D4EC8A")


def clean(value) -> str:
    return " ".join(str(value).replace("_x000D_", " / ").split()).strip() if value is not None else ""


def key(value) -> str:
    value = clean(value).upper().replace("2O TRAILER", "20 TRAILER")
    value = value.replace("COUPLEING", "COUPLING").replace("TRAILOR", "TRAILER")
    return re.sub(r"[^A-Z0-9]+", " ", value).strip()


def parse_dt(value) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    raw = clean(value).replace("\u202f", " ").replace("\xa0", " ")
    for pattern in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, pattern)
        except ValueError:
            pass
    raise ValueError(f"Unrecognised date/time {value!r}")


def money_cents(value) -> int:
    raw = re.sub(r"[^0-9.-]", "", clean(value))
    try:
        return int(round(float(raw or 0) * 100))
    except ValueError:
        return 0


def split_names(value) -> list[str]:
    return [clean(part) for part in re.split(r"\s*[;/]\s*|\s*\n\s*", clean(value)) if clean(part)]


def import_backup(source, db=None, *, branch=None, namespace=None, preserve_existing=True):
    source = Path(source).expanduser().resolve()
    if not source.is_file():
        raise ValueError(f"Workbook not found: {source}")
    if db:
        os.environ["A2Z_DATABASE"] = str(Path(db).expanduser().resolve())
    init_db()
    wb = load_workbook(source, read_only=True, data_only=True)
    conn = sqlite3.connect(database_path(), timeout=120)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=120000")
    report = Counter()
    errors = []
    try:
        if branch is None:
            raise ValueError("--branch is required so an import can never leak into another branch")
        branch_row = conn.execute(
            "SELECT id,name FROM branches WHERE is_active=1 AND (id=? OR lower(name)=lower(?))",
            (int(branch) if str(branch).isdigit() else -1, str(branch)),
        ).fetchone()
        if not branch_row:
            raise ValueError(f"Active branch not found: {branch}")
        branch_id = branch_row["id"]
        namespace = namespace or re.sub(r"[^a-z0-9]+", "-", branch_row["name"].lower()).strip("-")
        clients_added, clients_updated = import_clients(
            source, branch_id=branch_id, source_namespace=namespace,
            preserve_existing=preserve_existing,
        )
        report.update(clients_added=clients_added, clients_updated=clients_updated)
        admin = conn.execute(
            "SELECT id FROM users WHERE role='admin' AND branch_id=? ORDER BY is_super_admin,id LIMIT 1",
            (branch_id,),
        ).fetchone()
        admin_id = admin[0] if admin else None

        instructors = {key(r["name"]): r["id"] for r in conn.execute("SELECT id,name FROM instructors WHERE branch_id=?", (branch_id,))}
        machines = list(conn.execute("SELECT * FROM machines WHERE branch_id=?", (branch_id,)))
        machine_by_key = {key(r["machine_code"]): r for r in machines}

        def machine_for(text):
            target = key(text)
            exact = machine_by_key.get(target)
            if exact: return exact
            candidates = [r for mk, r in machine_by_key.items() if mk in target or target in mk]
            if candidates: return sorted(candidates, key=lambda r: len(key(r["machine_code"])), reverse=True)[0]
            return machines[0]

        # Restore the exact exported service catalogue alongside A2Z's grouped services.
        service_map = {}
        if "Services" in wb.sheetnames:
            ws = wb["Services"]; rows = ws.iter_rows(values_only=True)
            headers = [clean(v) for v in next(rows)]; ix = {h:i for i,h in enumerate(headers)}
            for pos, row in enumerate(rows):
                name = clean(row[ix["Service"]]);
                if not name: continue
                category = clean(row[ix["Group"]]) or "Imported"
                duration_raw = clean(row[ix["Duration"]]); match = re.search(r"\d+", duration_raw)
                duration = int(match.group()) if match else 30
                unavailable = {key(x) for x in re.split(r"[,;/]", clean(row[ix["Unavailable"]])) if clean(x)}
                weekdays = ",".join(str(n) for n, day in enumerate(("MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY","SUNDAY")) if day not in unavailable)
                conn.execute("""INSERT INTO services(branch_id,name,description,category,duration_minutes,price_cents,currency,color,available_weekdays,requires_approval,is_active)
                    VALUES (?,?,?, ?,?,?,'INR',?,?,0,1)
                    ON CONFLICT(branch_id,name) DO UPDATE SET category=excluded.category,duration_minutes=excluded.duration_minutes,price_cents=excluded.price_cents,available_weekdays=excluded.available_weekdays,is_active=1,updated_at=CURRENT_TIMESTAMP""",
                    (branch_id,name,"Imported from Smart Scheduling backup",category,duration,money_cents(row[ix["Price"]]),COLORS[pos%len(COLORS)],weekdays or ""))
                sid=conn.execute("SELECT id FROM services WHERE branch_id=? AND name=?",(branch_id,name)).fetchone()[0]
                service_map[key(name)] = sid
                if key(name) not in BREAK_NAMES:
                    conn.execute("INSERT OR IGNORE INTO service_machines(service_id,machine_id) VALUES (?,?)",(sid,machine_for(name)["id"]))
                for provider in split_names(row[ix["Providers"]]):
                    iid=instructors.get(key(provider))
                    if iid: conn.execute("INSERT OR IGNORE INTO service_instructors(service_id,instructor_id) VALUES (?,?)",(sid,iid))
                report["services"] += 1

        # Client lookup used by appointment rows (the export contains names, not IDs).
        by_phone, by_email, by_name = {}, {}, defaultdict(list)
        for r in conn.execute(
            "SELECT id,full_name,phone,email FROM users WHERE role='student' AND branch_id=?",
            (branch_id,),
        ):
            if key(r["phone"]): by_phone[key(r["phone"])] = r["id"]
            if key(r["email"]): by_email[key(r["email"])] = r["id"]
            by_name[key(r["full_name"])].append(r["id"])

        def client_id_for(name, phone, email, source_ref):
            uid = by_phone.get(key(phone)) or by_email.get(key(email))
            if not uid and len(by_name.get(key(name), [])) == 1: uid = by_name[key(name)][0]
            if uid: return uid
            username = "smart-appt-" + re.sub(r"[^a-z0-9]+", "-", source_ref.lower()).strip("-")
            conn.execute("""INSERT OR IGNORE INTO users(username,password_hash,role,full_name,email,phone,branch_id,is_active,login_enabled,must_change_password)
                VALUES (?,'!','student',?,?,?,?,1,0,0)""",(username,name or "Imported client",email or None,phone or None,branch_id))
            uid=conn.execute("SELECT id FROM users WHERE username=?",(username,)).fetchone()[0]
            conn.execute("INSERT OR IGNORE INTO client_profiles(user_id,reminders_enabled,preferred_channel,source_reference) VALUES (?,0,?,?)",(uid,"email" if email else "sms",source_ref))
            by_name[key(name)].append(uid)
            if key(phone): by_phone[key(phone)]=uid
            if key(email): by_email[key(email)]=uid
            report["appointment_clients_created"] += 1
            return uid

        conn.commit()
        conn.execute("BEGIN IMMEDIATE")
        for ws in wb.worksheets:
            if ws.title in {"Clients", "Services"}: continue
            sheet_key = key(ws.title.replace("...", ""))
            iid = instructors.get(sheet_key)
            if not iid:
                matches=[v for k,v in instructors.items() if k.startswith(sheet_key) or sheet_key.startswith(k[:20])]
                iid=matches[0] if len(set(matches))==1 else None
            if not iid:
                errors.append({"sheet":ws.title,"error":"No matching instructor"}); continue
            rows=ws.iter_rows(values_only=True); headers=[clean(v) for v in next(rows)]; ix={h:i for i,h in enumerate(headers)}
            for rn,row in enumerate(rows,2):
                if not any(v not in (None,"") for v in row): continue
                source_id=clean(row[ix["Id"]]) or f"row-{rn}"
                source_ref=f"smart:{namespace}:{ws.title}:{source_id}"
                try:
                    start=parse_dt(row[ix["Start"]]); end=parse_dt(row[ix["End"]])
                    if end <= start: raise ValueError("end is not after start")
                    client=clean(row[ix["Client"]]); phone=clean(row[ix["Phone"]]); email=clean(row[ix["Email"]])
                    service_text=clean(row[ix["Task/Services"]]); status_raw=key(row[ix["Status"]]); notes=clean(row[ix["Notes"]]); custom=clean(row[ix["Custom Fields"]])
                    if custom: notes=(notes+"\n" if notes else "")+"Custom fields: "+custom
                    duration=int((end-start).total_seconds()//60)
                    primary_service=split_names(service_text)[0] if split_names(service_text) else "Unspecified training"
                    m=machine_for(primary_service or client)
                    is_slot = "SLOT" in key(client) or (not client and duration >= 180 and key(primary_service) not in BREAK_NAMES)
                    is_busy = key(primary_service) in BREAK_NAMES or key(client) in BREAK_NAMES
                    values=(iid,m["id"],branch_id,start.date().isoformat(),start.strftime("%H:%M"),end.strftime("%H:%M"))
                    if is_slot:
                        if preserve_existing and conn.execute("SELECT 1 FROM booking_slots WHERE source_reference=?", (source_ref,)).fetchone():
                            report["slots_skipped_existing"] += 1
                            continue
                        conn.execute("""INSERT INTO booking_slots(instructor_id,machine_id,branch_id,target_date,start_time,end_time,notes,series_id,repeat_rule,created_by,source_reference)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(source_reference) DO UPDATE SET instructor_id=excluded.instructor_id,machine_id=excluded.machine_id,target_date=excluded.target_date,start_time=excluded.start_time,end_time=excluded.end_time,notes=excluded.notes,updated_at=CURRENT_TIMESTAMP""",
                            (*values, notes or client, clean(row[ix["Series-Id"]]) or None, None, admin_id, source_ref))
                        report["slots"] += 1
                    elif is_busy:
                        if preserve_existing and conn.execute("SELECT 1 FROM instructor_time_off WHERE source_reference=?", (source_ref,)).fetchone():
                            report["busy_skipped_existing"] += 1
                            continue
                        conn.execute("""INSERT INTO instructor_time_off(instructor_id,target_date,start_time,end_time,reason,notes,series_id,created_by,source_reference)
                            VALUES (?,?,?,?,?,?,?,?,?) ON CONFLICT(source_reference) DO UPDATE SET instructor_id=excluded.instructor_id,target_date=excluded.target_date,start_time=excluded.start_time,end_time=excluded.end_time,reason=excluded.reason,notes=excluded.notes,updated_at=CURRENT_TIMESTAMP""",
                            (iid,start.date().isoformat(),start.strftime("%H:%M"),end.strftime("%H:%M"),primary_service or client,notes,clean(row[ix["Series-Id"]]) or None,admin_id,source_ref))
                        report["busy"] += 1
                    else:
                        if preserve_existing and conn.execute("SELECT 1 FROM bookings WHERE source_reference=?", (source_ref,)).fetchone():
                            report["appointments_skipped_existing"] += 1
                            continue
                        sid=service_map.get(key(primary_service))
                        uid=client_id_for(client,phone,email,source_ref)
                        mapped=STATUS_MAP.get(status_raw,"Pending")
                        cur=conn.execute("""INSERT INTO bookings(student_name,mobile_number,machine_id,instructor_id,branch_id,target_date,start_time,end_time,validation_status,student_user_id,notes,service_id,service_name,service_price_cents,currency,buffer_before_minutes,buffer_after_minutes,series_id,allow_double_booking,source_reference)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'INR',0,0,?,1,?) ON CONFLICT(source_reference) DO UPDATE SET student_name=excluded.student_name,mobile_number=excluded.mobile_number,machine_id=excluded.machine_id,instructor_id=excluded.instructor_id,target_date=excluded.target_date,start_time=excluded.start_time,end_time=excluded.end_time,validation_status=excluded.validation_status,student_user_id=excluded.student_user_id,notes=excluded.notes,service_id=excluded.service_id,service_name=excluded.service_name,service_price_cents=excluded.service_price_cents,updated_at=CURRENT_TIMESTAMP RETURNING id""",
                            (client or "Imported appointment",phone,m["id"],iid,branch_id,start.date().isoformat(),start.strftime("%H:%M"),end.strftime("%H:%M"),mapped,uid,notes,sid,service_text,money_cents(row[ix["Price"]]),clean(row[ix["Series-Id"]]) or None,source_ref))
                        bid=cur.fetchone()[0]
                        conn.execute("DELETE FROM booking_services WHERE booking_id=?",(bid,))
                        for order,name in enumerate(split_names(service_text) or ["Unspecified training"]):
                            service_id=service_map.get(key(name)); mins=duration if order==0 else 0
                            conn.execute("INSERT INTO booking_services(booking_id,service_id,service_name,duration_minutes,price_cents,currency,sort_order) VALUES (?,?,?,?,?,'INR',?)",(bid,service_id,name,mins,money_cents(row[ix["Price"]]) if order==0 else 0,order))
                        report["appointments"] += 1; report["status_"+mapped] += 1
                except Exception as exc:
                    errors.append({"sheet":ws.title,"row":rn,"id":source_id,"error":str(exc)})
                    report["errors"] += 1
        conn.commit()
    except Exception:
        conn.rollback(); raise
    finally:
        conn.close(); wb.close()
    return dict(report), errors


def main():
    parser=argparse.ArgumentParser(); parser.add_argument("workbook"); parser.add_argument("--database"); parser.add_argument("--report")
    parser.add_argument("--branch", required=True, help="Target branch ID or exact branch name")
    parser.add_argument("--namespace", help="Stable source namespace; defaults to branch name")
    parser.add_argument("--update-existing", action="store_true", help="Update matching source records instead of preserving them")
    args=parser.parse_args(); report,errors=import_backup(
        args.workbook,args.database,branch=args.branch,namespace=args.namespace,
        preserve_existing=not args.update_existing,
    )
    result={"database":str(database_path()),"counts":report,"errors":errors}
    if args.report: Path(args.report).write_text(json.dumps(result,indent=2),encoding="utf-8")
    print(json.dumps(result,indent=2))
    if errors: raise SystemExit(2)


if __name__ == "__main__": main()
