"""Safely import a Smart Scheduling export into the A2Z Driving School portal.

This importer never writes to the Heavy Equipment branch.  It imports every
client, instructor sheet, service and resource from the workbook, but only
calendar records whose start date/time is still in the future when it runs.
It is idempotent: running it again updates imported records instead of creating
duplicates.

Examples (inside the Coolify application terminal)::

    python import_driving_school_smart_backup.py /data/import/backup.xlsx
    python import_driving_school_smart_backup.py /data/import/backup.xlsx --apply
"""

from __future__ import annotations

import argparse
import os
import re
import sqlite3
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from database import database_path, init_db, seed_portal_accounts


PORTAL_NAME = os.environ.get("A2Z_DRIVING_SCHOOL_PORTAL_NAME", "A2Z Driving School")
BREAK_NAMES = {"BREAKFAST", "LUNCH BREAK", "TEA BREAK", "NO BOOKING", "LEAVE"}
STATUS_MAP = {
    "": "Approved", "NONE": "Approved", "CONFIRMED": "Approved",
    "NOT CONFIRMED": "Pending", "PENDING": "Pending",
    "COMPLETED": "Completed", "CANCELLED": "Cancelled",
    "NO-SHOW": "No-show", "NO SHOW": "No-show", "RESCHEDULED": "Rescheduled",
    "RUNNING LATE": "Running Late", "ARRIVED": "Arrived",
}
COLOURS = ("#E6C8F5", "#7CC5E3", "#F5E39B", "#B9E5BF", "#F1DCC0", "#D4EC8A")


def clean(value) -> str:
    return " ".join(str(value).replace("_x000D_", " / ").split()).strip() if value is not None else ""


def key(value) -> str:
    value = clean(value).upper().replace("2O TRAILER", "20 TRAILER")
    value = value.replace("COUPLEING", "COUPLING").replace("TRAILOR", "TRAILER")
    return re.sub(r"[^A-Z0-9]+", " ", value).strip()


def slug(value) -> str:
    return re.sub(r"[^a-z0-9]+", "-", clean(value).lower()).strip("-") or "unknown"


def parse_dt(value) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    raw = clean(value).replace("\u202f", " ").replace("\xa0", " ")
    for pattern in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, pattern)
        except ValueError:
            pass
    raise ValueError(f"Unrecognised date/time {value!r}")


def money_cents(value) -> int:
    try:
        return int(round(float(re.sub(r"[^0-9.-]", "", clean(value)) or 0) * 100))
    except ValueError:
        return 0


def split_services(value) -> list[str]:
    return [clean(part) for part in re.split(r"\s*[;/]\s*|\s*\n\s*", clean(value)) if clean(part)]


def headers_and_rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    return {name: index for index, name in enumerate(headers) if name}, rows


def make_backup(db_path: Path) -> Path:
    backup_dir = Path(os.environ.get("A2Z_BACKUP_DIR", db_path.parent / "backups"))
    backup_dir.mkdir(parents=True, exist_ok=True)
    destination = backup_dir / f"a2z-before-driving-school-import-{datetime.now():%Y%m%d-%H%M%S}.db"
    source = sqlite3.connect(db_path)
    target = sqlite3.connect(destination)
    try:
        source.backup(target)
    finally:
        target.close()
        source.close()
    return destination


def is_busy(client: str, service: str) -> bool:
    return key(client) in BREAK_NAMES or key(service) in BREAK_NAMES or "NO BOOKING" in key(client)


def collect_future_counts(workbook, cutoff: datetime) -> Counter:
    counts = Counter()
    for sheet in workbook.worksheets:
        if sheet.title in {"Clients", "Services"}:
            continue
        columns, rows = headers_and_rows(sheet)
        if "Start" not in columns:
            continue
        for row in rows:
            try:
                start = parse_dt(row[columns["Start"]])
            except (ValueError, IndexError):
                continue
            if start >= cutoff:
                counts["future_calendar_records"] += 1
    return counts


def import_workbook(source: Path, *, apply: bool) -> tuple[Counter, Path | None]:
    source = source.expanduser().resolve()
    if not source.is_file():
        raise ValueError(f"Workbook not found: {source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    if "Clients" not in workbook.sheetnames or "Services" not in workbook.sheetnames:
        raise ValueError("This does not look like a Smart Scheduling backup (Clients and Services sheets are required).")

    cutoff = datetime.now().replace(second=0, microsecond=0)
    counts = Counter(clients=len(range(2, workbook["Clients"].max_row + 1)))
    counts["instructors"] = len([ws for ws in workbook.worksheets if ws.title not in {"Clients", "Services"}])
    counts["services"] = max(0, workbook["Services"].max_row - 1)
    counts.update(collect_future_counts(workbook, cutoff))
    if not apply:
        workbook.close()
        return counts, None

    init_db()
    seed_portal_accounts()
    db_path = database_path()
    backup = make_backup(db_path)
    conn = sqlite3.connect(db_path, timeout=120)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=120000")
    try:
        branch = conn.execute("SELECT id FROM branches WHERE name=? AND is_active=1", (PORTAL_NAME,)).fetchone()
        if not branch:
            raise RuntimeError(f"Driving School portal {PORTAL_NAME!r} does not exist. Add the portal environment variables and redeploy first.")
        branch_id = branch["id"]
        admin = conn.execute("SELECT id FROM users WHERE username=?", (os.environ.get("A2Z_DRIVING_SCHOOL_ADMIN_USERNAME", "admin_drivingschool"),)).fetchone()
        admin_id = admin["id"] if admin else None
        conn.execute("BEGIN IMMEDIATE")

        # The workbook tab names are the Driving School instructors.
        instructors: dict[str, int] = {}
        for sheet in workbook.worksheets:
            if sheet.title in {"Clients", "Services"}:
                continue
            name = clean(sheet.title)
            conn.execute(
                """INSERT INTO instructors(name,branch_id,specialty,is_active,verification_status,verified_at,verified_by)
                   VALUES (?,?,'Imported from Smart Scheduling',1,'verified',CURRENT_TIMESTAMP,?)
                   ON CONFLICT(name,branch_id) DO UPDATE SET is_active=1,verification_status='verified',updated_at=CURRENT_TIMESTAMP""",
                (name, branch_id, admin_id),
            )
            instructors[key(name)] = conn.execute("SELECT id FROM instructors WHERE name=? AND branch_id=?", (name, branch_id)).fetchone()[0]

        # Create a Driving School resource for each non-break service. Prefix keeps
        # resources separate from similarly named Heavy Equipment resources.
        services: dict[str, tuple[int, int]] = {}
        columns, rows = headers_and_rows(workbook["Services"])
        for position, row in enumerate(rows):
            name = clean(row[columns.get("Service", 0)])
            if not name:
                continue
            group = clean(row[columns.get("Group", 1)]) or "Driving School"
            duration = int(re.search(r"\d+", clean(row[columns.get("Duration", 2)]) or "15").group())
            price = money_cents(row[columns.get("Price", 3)])
            unavailable = {key(day) for day in re.split(r"[,;/]", clean(row[columns.get("Unavailable", 5)])) if clean(day)}
            weekdays = ",".join(str(number) for number, day in enumerate(("MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY")) if day not in unavailable)
            conn.execute(
                """INSERT INTO services(branch_id,name,description,category,duration_minutes,price_cents,currency,color,available_weekdays,requires_approval,is_active)
                   VALUES (?,?,?, ?,?,?,'INR',?,?,0,1)
                   ON CONFLICT(branch_id,name) DO UPDATE SET category=excluded.category,duration_minutes=excluded.duration_minutes,price_cents=excluded.price_cents,color=excluded.color,available_weekdays=excluded.available_weekdays,is_active=1,updated_at=CURRENT_TIMESTAMP""",
                (branch_id, name, "Imported from Smart Scheduling backup", group, duration, price, COLOURS[position % len(COLOURS)], weekdays),
            )
            service_id = conn.execute("SELECT id FROM services WHERE branch_id=? AND name=?", (branch_id, name)).fetchone()[0]
            resource_name = f"Driving School · {name}"
            conn.execute(
                """INSERT INTO machines(machine_code,category,location,branch_id,is_active)
                   VALUES (?,?,?, ?,1)
                   ON CONFLICT(machine_code) DO UPDATE SET category=excluded.category,location=excluded.location,branch_id=excluded.branch_id,is_active=1""",
                (resource_name, group, "Driving School", branch_id),
            )
            machine_id = conn.execute("SELECT id FROM machines WHERE machine_code=?", (resource_name,)).fetchone()[0]
            services[key(name)] = (service_id, machine_id)
            if key(name) not in BREAK_NAMES:
                conn.execute("INSERT OR IGNORE INTO service_machines(service_id,machine_id) VALUES (?,?)", (service_id, machine_id))
            for provider in split_services(row[columns.get("Providers", 6)]):
                instructor_id = instructors.get(key(provider))
                if instructor_id:
                    conn.execute("INSERT OR IGNORE INTO service_instructors(service_id,instructor_id) VALUES (?,?)", (service_id, instructor_id))

        # Every Smart Scheduling client receives a branch-isolated identity.
        client_by_phone: dict[str, int] = {}
        client_by_name: dict[str, list[int]] = {}
        client_columns, client_rows = headers_and_rows(workbook["Clients"])
        for row_number, row in enumerate(client_rows, 2):
            source_id = clean(row[client_columns.get("Id", 0)]) or f"row-{row_number}"
            username = f"smart-driving-client-{slug(source_id)}"
            first = clean(row[client_columns.get("First Name", 1)])
            last = clean(row[client_columns.get("Last Name", 2)])
            full_name = " ".join(part for part in (first, last) if part) or f"Client {source_id}"
            email = clean(row[client_columns.get("Email", 3)]) or None
            phone = clean(row[client_columns.get("Phone", 5)]) or None
            existing = conn.execute("SELECT id FROM users WHERE lower(username)=lower(?)", (username,)).fetchone()
            if existing:
                user_id = existing["id"]
                conn.execute("UPDATE users SET full_name=?,email=?,phone=?,branch_id=?,is_active=1,login_enabled=0,must_change_password=0,updated_at=CURRENT_TIMESTAMP WHERE id=?", (full_name, email, phone, branch_id, user_id))
                counts["clients_updated"] += 1
            else:
                user_id = conn.execute("INSERT INTO users(username,password_hash,role,full_name,email,phone,branch_id,is_active,login_enabled,must_change_password) VALUES (?,'!','student',?,?,?,?,1,0,0)", (username, full_name, email, phone, branch_id)).lastrowid
                counts["clients_added"] += 1
            conn.execute(
                """INSERT INTO client_profiles(user_id,secondary_phone,secondary_email,birthday,gender,zip_code,city,street,internal_notes,reminders_enabled,preferred_channel,source_reference)
                   VALUES (?,?,?,?,?,?,?,?,?,1,?,?)
                   ON CONFLICT(user_id) DO UPDATE SET secondary_phone=excluded.secondary_phone,secondary_email=excluded.secondary_email,zip_code=excluded.zip_code,city=excluded.city,street=excluded.street,internal_notes=excluded.internal_notes,source_reference=excluded.source_reference,updated_at=CURRENT_TIMESTAMP""",
                (user_id, clean(row[client_columns.get("Secondary Phone", 6)]) or None, clean(row[client_columns.get("Secondary Email", 4)]) or None, None, None, clean(row[client_columns.get("Zip Code", 11)]) or None, clean(row[client_columns.get("City", 10)]) or None, clean(row[client_columns.get("Street", 9)]) or None, clean(row[client_columns.get("Notes", 13)]) or None, "email" if email else "sms", f"smart-driving-client:{source_id}"),
            )
            if phone:
                client_by_phone[key(phone)] = user_id
            client_by_name.setdefault(key(full_name), []).append(user_id)

        def client_for(name: str, phone: str, source_reference: str) -> int:
            user_id = client_by_phone.get(key(phone))
            if not user_id and len(client_by_name.get(key(name), [])) == 1:
                user_id = client_by_name[key(name)][0]
            if user_id:
                return user_id
            username = f"smart-driving-appointment-{slug(source_reference)}"
            conn.execute("INSERT OR IGNORE INTO users(username,password_hash,role,full_name,phone,branch_id,is_active,login_enabled,must_change_password) VALUES (?,'!','student',?,?,?,1,0,0)", (username, name or "Imported client", phone or None, branch_id))
            user_id = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()[0]
            conn.execute("INSERT OR IGNORE INTO client_profiles(user_id,reminders_enabled,preferred_channel,source_reference) VALUES (?,0,'sms',?)", (user_id, source_reference))
            counts["appointment_clients_added"] += 1
            return user_id

        fallback = next(iter(services.values()), None)
        if not fallback:
            raise RuntimeError("The Services sheet has no usable service rows.")
        for sheet in workbook.worksheets:
            if sheet.title in {"Clients", "Services"}:
                continue
            instructor_id = instructors[key(sheet.title)]
            columns, rows = headers_and_rows(sheet)
            for row_number, row in enumerate(rows, 2):
                try:
                    start = parse_dt(row[columns["Start"]])
                    end = parse_dt(row[columns["End"]])
                    if end <= start or start < cutoff:
                        continue
                    source_id = clean(row[columns.get("Id", 0)]) or f"row-{row_number}"
                    source_reference = f"smart-driving:{sheet.title}:{source_id}"
                    client = clean(row[columns.get("Client", 4)])
                    phone = clean(row[columns.get("Phone", 6)])
                    service_text = clean(row[columns.get("Task/Services", 9)])
                    status = STATUS_MAP.get(key(row[columns.get("Status", 10)]), "Pending")
                    notes = clean(row[columns.get("Notes", 11)])
                    primary = split_services(service_text)[0] if split_services(service_text) else "Unspecified training"
                    service_id, machine_id = services.get(key(primary), fallback)
                    if is_busy(client, primary) or not client:
                        conn.execute(
                            """INSERT INTO instructor_time_off(instructor_id,target_date,start_time,end_time,reason,notes,series_id,created_by,source_reference)
                               VALUES (?,?,?,?,?,?,?,?,?)
                               ON CONFLICT(source_reference) DO UPDATE SET instructor_id=excluded.instructor_id,target_date=excluded.target_date,start_time=excluded.start_time,end_time=excluded.end_time,reason=excluded.reason,notes=excluded.notes,updated_at=CURRENT_TIMESTAMP""",
                            (instructor_id, start.date().isoformat(), start.strftime("%H:%M"), end.strftime("%H:%M"), primary or client or "No booking", notes, clean(row[columns.get("Series-Id", 8)]) or None, admin_id, source_reference),
                        )
                        counts["future_blocks"] += 1
                        continue
                    client_id = client_for(client, phone, source_reference)
                    cursor = conn.execute(
                        """INSERT INTO bookings(student_name,mobile_number,machine_id,instructor_id,branch_id,target_date,start_time,end_time,validation_status,student_user_id,notes,service_id,service_name,service_price_cents,currency,buffer_before_minutes,buffer_after_minutes,series_id,allow_double_booking,source_reference)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'INR',0,0,?,1,?)
                           ON CONFLICT(source_reference) DO UPDATE SET student_name=excluded.student_name,mobile_number=excluded.mobile_number,machine_id=excluded.machine_id,instructor_id=excluded.instructor_id,target_date=excluded.target_date,start_time=excluded.start_time,end_time=excluded.end_time,validation_status=excluded.validation_status,student_user_id=excluded.student_user_id,notes=excluded.notes,service_id=excluded.service_id,service_name=excluded.service_name,service_price_cents=excluded.service_price_cents,updated_at=CURRENT_TIMESTAMP RETURNING id""",
                        (client, phone, machine_id, instructor_id, branch_id, start.date().isoformat(), start.strftime("%H:%M"), end.strftime("%H:%M"), status, client_id, notes, service_id, service_text or primary, money_cents(row[columns.get("Price", 7)]), clean(row[columns.get("Series-Id", 8)]) or None, source_reference),
                    )
                    booking_id = cursor.fetchone()[0]
                    conn.execute("DELETE FROM booking_services WHERE booking_id=?", (booking_id,))
                    for order, service_name in enumerate(split_services(service_text) or [primary]):
                        item_service_id = services.get(key(service_name), (None, None))[0]
                        duration = max(0, int((end - start).total_seconds() // 60)) if order == 0 else 0
                        conn.execute("INSERT INTO booking_services(booking_id,service_id,service_name,duration_minutes,price_cents,currency,sort_order) VALUES (?,?,?,?,?,'INR',?)", (booking_id, item_service_id, service_name, duration, money_cents(row[columns.get("Price", 7)]) if order == 0 else 0, order))
                    counts["future_appointments"] += 1
                except (IndexError, KeyError, ValueError) as exc:
                    counts["skipped_invalid_rows"] += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
        workbook.close()
    return counts, backup


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Smart Scheduling backup XLSX")
    parser.add_argument("--apply", action="store_true", help="Write the import. Omit for a safe preview.")
    args = parser.parse_args()
    counts, backup = import_workbook(Path(args.workbook), apply=args.apply)
    print("Import preview:" if not args.apply else "Driving School import complete:")
    for name, value in sorted(counts.items()):
        print(f"  {name.replace('_', ' ')}: {value}")
    if backup:
        print(f"Database backup created: {backup}")
    elif not args.apply:
        print("No data was changed. Re-run with --apply after checking this preview.")


if __name__ == "__main__":
    main()
