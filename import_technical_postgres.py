"""Idempotently import the Technical Smart Scheduling export into PostgreSQL.

This utility never truncates tables.  It requires the exact target branch and
uses namespaced source references so reruns skip records already imported.
Preview is the default; pass --apply only after a verified pg_dump exists.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from database import get_db, postgres_url


BREAK_NAMES = {"BREAK", "BREAKFAST", "TEA BREAK", "TEA TIME", "TEATIME", "LUNCH", "LUNCH BREAK", "NO BOOKING", "LEAVE"}
STATUS_MAP = {
    "": "Approved", "NONE": "Approved", "CONFIRMED": "Approved",
    "NOT CONFIRMED": "Pending", "PENDING": "Pending",
    "COMPLETED": "Completed", "CANCELLED": "Cancelled",
    "NO-SHOW": "No-show", "NO SHOW": "No-show",
    "RESCHEDULED": "Rescheduled", "RUNNING LATE": "Running Late",
    "ARRIVED": "Arrived",
}
COLORS = ("#E6C8F5", "#7CC5E3", "#F5E39B", "#B9E5BF", "#F1DCC0", "#D4EC8A")


def clean(value) -> str:
    return " ".join(str(value).replace("_x000D_", " / ").replace("\u202f", " ").replace("\xa0", " ").split()).strip() if value is not None else ""


def key(value) -> str:
    value = clean(value).upper().replace("2O TRAILER", "20 TRAILER")
    value = value.replace("COUPLEING", "COUPLING").replace("TRAILOR", "TRAILER")
    return re.sub(r"[^A-Z0-9]+", " ", value).strip()


def slug(value) -> str:
    return re.sub(r"[^a-z0-9]+", "-", clean(value).lower()).strip("-")


def parse_dt(value) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    raw = clean(value)
    for pattern in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, pattern)
        except ValueError:
            pass
    raise ValueError(f"Unrecognised date/time {value!r}")


def money_cents(value) -> int:
    raw = re.sub(r"[^0-9.-]", "", clean(value))
    try:
        return int(round(float(raw or 0) * 100))
    except ValueError:
        return 0


def split_names(value) -> list[str]:
    return [clean(part) for part in re.split(r"\s*[;/]\s*|\s*\n\s*", clean(value)) if clean(part)]


def headers_and_rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    return {name: index for index, name in enumerate(headers)}, rows


def workbook_totals(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            "clients": max(0, wb["Clients"].max_row - 1) if "Clients" in wb.sheetnames else 0,
            "services": max(0, wb["Services"].max_row - 1) if "Services" in wb.sheetnames else 0,
            "calendar_sheets": len([name for name in wb.sheetnames if name not in {"Clients", "Services"}]),
            "calendar_rows": sum(max(0, wb[name].max_row - 1) for name in wb.sheetnames if name not in {"Clients", "Services"}),
        }
    finally:
        wb.close()


def import_clients(conn, workbook_path: Path, branch_id: int, report: Counter) -> None:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = wb["Clients"] if "Clients" in wb.sheetnames else wb.active
        columns, rows = headers_and_rows(sheet)
        for row_number, row in enumerate(rows, 2):
            source_id = clean(row[columns["Id"]]) or f"row-{row_number}"
            username = f"smart-client-technical-{slug(source_id) or row_number}"
            existing = conn.execute("SELECT id FROM users WHERE lower(username)=lower(?)", (username,)).fetchone()
            if existing:
                report["clients_skipped_existing"] += 1
                continue
            first = clean(row[columns["First Name"]])
            last = clean(row[columns["Last Name"]])
            full_name = " ".join(value for value in (first, last) if value) or f"Client {source_id}"
            email = clean(row[columns["Email"]]) or None
            phone = clean(row[columns["Phone"]]) or None
            user_id = conn.execute_insert(
                """INSERT INTO users
                   (username,password_hash,role,full_name,email,phone,branch_id,
                    is_active,login_enabled,must_change_password)
                   VALUES (?,'!','student',?,?,?,?,1,0,0)""",
                (username, full_name, email, phone, branch_id),
            ).lastrowid
            gender = clean(row[columns["Gender"]]).lower()
            gender = gender if gender in {"male", "female", "other"} else None
            conn.execute(
                """INSERT INTO client_profiles
                   (user_id,secondary_phone,secondary_email,gender,zip_code,city,
                    street,internal_notes,reminders_enabled,preferred_channel,
                    source_reference)
                   VALUES (?,?,?,?,?,?,?,?,1,?,?)""",
                (
                    user_id,
                    clean(row[columns["Secondary Phone"]]) or None,
                    clean(row[columns["Secondary Email"]]) or None,
                    gender,
                    clean(row[columns["Zip Code"]]) or None,
                    clean(row[columns["City"]]) or None,
                    clean(row[columns["Street"]]) or None,
                    clean(row[columns["Notes"]]) or None,
                    "email" if email else "sms",
                    f"smart:technical:client:{source_id}",
                ),
            )
            report["clients_added"] += 1
    finally:
        wb.close()


def apply_import(backup_path: Path, clients_path: Path, branch_name: str) -> tuple[dict, list]:
    report = Counter()
    errors = []
    wb = load_workbook(backup_path, read_only=True, data_only=True)
    try:
        with get_db() as conn:
            branch = conn.execute("SELECT id,name FROM branches WHERE lower(name)=lower(?) AND is_active=1", (branch_name,)).fetchone()
            if not branch:
                raise ValueError(f"Active target branch not found: {branch_name}")
            branch_id = int(branch["id"])
            if branch_id in {1, 120}:
                raise ValueError("Safety stop: the Technical importer cannot target Heavy or Driving School")
            admin = conn.execute("SELECT id FROM users WHERE role='admin' AND branch_id=? ORDER BY is_super_admin,id LIMIT 1", (branch_id,)).fetchone()
            admin_id = admin["id"] if admin else None

            instructor_names = [ws.title.strip() for ws in wb.worksheets if ws.title not in {"Clients", "Services"}]
            if "Services" in wb.sheetnames:
                columns, rows = headers_and_rows(wb["Services"])
                for row in rows:
                    for provider in split_names(row[columns["Providers"]]):
                        if provider not in instructor_names:
                            instructor_names.append(provider)
            instructors = {}
            for name in instructor_names:
                conn.execute(
                    """INSERT OR IGNORE INTO instructors
                       (name,branch_id,specialty,is_active,verification_status)
                       VALUES (?,?,'Technical training',1,'verified')""",
                    (name, branch_id),
                )
                item = conn.execute("SELECT id FROM instructors WHERE branch_id=? AND lower(name)=lower(?)", (branch_id, name)).fetchone()
                instructors[key(name)] = item["id"]
            report["instructors_ready"] = len(instructors)

            service_names = []
            if "Services" in wb.sheetnames:
                columns, rows = headers_and_rows(wb["Services"])
                service_names = [clean(row[columns["Service"]]) for row in rows if clean(row[columns["Service"]])]
            for name in service_names + ["GENERAL"]:
                machine_code = "TECH-" + key(name)
                conn.execute(
                    """INSERT OR IGNORE INTO machines
                       (machine_code,category,location,branch_id,is_active)
                       VALUES (?,'Technical','Technical',?,1)""",
                    (machine_code, branch_id),
                )

            import_clients(conn, backup_path, branch_id, report)
            import_clients(conn, clients_path, branch_id, report)

            by_phone, by_email, by_name = {}, {}, defaultdict(list)
            for item in conn.execute("SELECT id,full_name,phone,email FROM users WHERE role='student' AND branch_id=?", (branch_id,)):
                if key(item["phone"]): by_phone[key(item["phone"])] = item["id"]
                if key(item["email"]): by_email[key(item["email"])] = item["id"]
                by_name[key(item["full_name"])].append(item["id"])

            def client_for(name, phone, email, source_ref):
                user_id = by_phone.get(key(phone)) or by_email.get(key(email))
                if not user_id and len(by_name.get(key(name), [])) == 1:
                    user_id = by_name[key(name)][0]
                if user_id:
                    return user_id
                username = "smart-appt-technical-" + slug(source_ref)
                existing = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
                if existing:
                    return existing["id"]
                user_id = conn.execute_insert(
                    """INSERT INTO users
                       (username,password_hash,role,full_name,email,phone,branch_id,
                        is_active,login_enabled,must_change_password)
                       VALUES (?,'!','student',?,?,?,?,1,0,0)""",
                    (username, name or "Imported client", email or None, phone or None, branch_id),
                ).lastrowid
                conn.execute(
                    """INSERT INTO client_profiles
                       (user_id,reminders_enabled,preferred_channel,source_reference)
                       VALUES (?,0,?,?)""",
                    (user_id, "email" if email else "sms", source_ref),
                )
                by_name[key(name)].append(user_id)
                if key(phone): by_phone[key(phone)] = user_id
                if key(email): by_email[key(email)] = user_id
                report["appointment_clients_added"] += 1
                return user_id

            machines = list(conn.execute("SELECT * FROM machines WHERE branch_id=? AND is_active=1", (branch_id,)))
            machine_by_key = {key(item["machine_code"]): item for item in machines}

            def machine_for(text):
                target = key(text)
                candidates = [item for machine_key, item in machine_by_key.items() if target and (target in machine_key or machine_key in target)]
                if candidates:
                    return sorted(candidates, key=lambda item: len(key(item["machine_code"])), reverse=True)[0]
                return next(item for item in machines if item["machine_code"] == "TECH-GENERAL")

            service_map = {}
            if "Services" in wb.sheetnames:
                columns, rows = headers_and_rows(wb["Services"])
                for position, row in enumerate(rows):
                    name = clean(row[columns["Service"]])
                    if not name: continue
                    duration_match = re.search(r"\d+", clean(row[columns["Duration"]]))
                    duration = int(duration_match.group()) if duration_match else 30
                    category = clean(row[columns["Group"]]) or "Imported"
                    conn.execute(
                        """INSERT OR IGNORE INTO services
                           (branch_id,name,description,category,duration_minutes,
                            price_cents,currency,buffer_before_minutes,
                            buffer_after_minutes,color,available_weekdays,
                            requires_approval,is_active)
                           VALUES (?,?,?,?,?,?,'INR',0,0,?,'0,1,2,3,4,5,6',0,1)""",
                        (branch_id, name, "Imported Technical service", category, duration, money_cents(row[columns["Price"]]), COLORS[position % len(COLORS)]),
                    )
                    service = conn.execute("SELECT id FROM services WHERE branch_id=? AND lower(name)=lower(?)", (branch_id, name)).fetchone()
                    service_map[key(name)] = service["id"]
                    if key(name) not in BREAK_NAMES:
                        conn.execute("INSERT OR IGNORE INTO service_machines(service_id,machine_id) VALUES (?,?)", (service["id"], machine_for(name)["id"]))
                    for provider in split_names(row[columns["Providers"]]):
                        instructor_id = instructors.get(key(provider))
                        if instructor_id:
                            conn.execute("INSERT OR IGNORE INTO service_instructors(service_id,instructor_id) VALUES (?,?)", (service["id"], instructor_id))
                    report["services_ready"] += 1

            for sheet in wb.worksheets:
                if sheet.title in {"Clients", "Services"}: continue
                instructor_id = instructors.get(key(sheet.title))
                if not instructor_id:
                    errors.append({"sheet": sheet.title, "error": "No matching instructor"})
                    continue
                columns, rows = headers_and_rows(sheet)
                for row_number, row in enumerate(rows, 2):
                    if not any(value not in (None, "") for value in row): continue
                    source_id = clean(row[columns["Id"]]) or f"row-{row_number}"
                    source_ref = f"smart:technical:{sheet.title}:{source_id}"
                    try:
                        start = parse_dt(row[columns["Start"]]); end = parse_dt(row[columns["End"]])
                        if end <= start: raise ValueError("end is not after start")
                        client = clean(row[columns["Client"]]); phone = clean(row[columns["Phone"]]); email = clean(row[columns["Email"]])
                        service_text = clean(row[columns["Task/Services"]]); notes = clean(row[columns["Notes"]])
                        custom = clean(row[columns["Custom Fields"]])
                        if custom: notes = (notes + "\n" if notes else "") + "Custom fields: " + custom
                        duration = int((end - start).total_seconds() // 60)
                        names = split_names(service_text)
                        primary = names[0] if names else "Unspecified training"
                        machine = machine_for(primary or client)
                        is_slot = "SLOT" in key(client) or (not client and duration >= 180 and key(primary) not in BREAK_NAMES)
                        is_busy = key(primary) in BREAK_NAMES or key(client) in BREAK_NAMES
                        if is_slot:
                            if conn.execute("SELECT 1 FROM booking_slots WHERE source_reference=?", (source_ref,)).fetchone():
                                report["slots_skipped_existing"] += 1; continue
                            conn.execute(
                                """INSERT INTO booking_slots
                                   (instructor_id,machine_id,branch_id,target_date,
                                    start_time,end_time,notes,series_id,created_by,
                                    source_reference)
                                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                                (instructor_id, machine["id"], branch_id, start.date().isoformat(), start.strftime("%H:%M"), end.strftime("%H:%M"), notes or client, clean(row[columns["Series-Id"]]) or None, admin_id, source_ref),
                            )
                            report["slots_added"] += 1
                        elif is_busy:
                            if conn.execute("SELECT 1 FROM instructor_time_off WHERE source_reference=?", (source_ref,)).fetchone():
                                report["busy_skipped_existing"] += 1; continue
                            conn.execute(
                                """INSERT INTO instructor_time_off
                                   (instructor_id,target_date,start_time,end_time,
                                    reason,notes,series_id,created_by,source_reference)
                                   VALUES (?,?,?,?,?,?,?,?,?)""",
                                (instructor_id, start.date().isoformat(), start.strftime("%H:%M"), end.strftime("%H:%M"), primary or client, notes, clean(row[columns["Series-Id"]]) or None, admin_id, source_ref),
                            )
                            report["busy_added"] += 1
                        else:
                            if conn.execute("SELECT 1 FROM bookings WHERE source_reference=?", (source_ref,)).fetchone():
                                report["appointments_skipped_existing"] += 1; continue
                            user_id = client_for(client, phone, email, source_ref)
                            status = STATUS_MAP.get(key(row[columns["Status"]]), "Pending")
                            service_id = service_map.get(key(primary))
                            booking_id = conn.execute_insert(
                                """INSERT INTO bookings
                                   (student_name,mobile_number,machine_id,
                                    instructor_id,branch_id,target_date,start_time,
                                    end_time,validation_status,student_user_id,
                                    notes,service_id,service_name,
                                    service_price_cents,currency,
                                    buffer_before_minutes,buffer_after_minutes,
                                    series_id,allow_double_booking,source_reference)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'INR',0,0,?,1,?)""",
                                (client or "Imported appointment", phone, machine["id"], instructor_id, branch_id, start.date().isoformat(), start.strftime("%H:%M"), end.strftime("%H:%M"), status, user_id, notes, service_id, service_text, money_cents(row[columns["Price"]]), clean(row[columns["Series-Id"]]) or None, source_ref),
                            ).lastrowid
                            for order, name in enumerate(names or ["Unspecified training"]):
                                conn.execute(
                                    """INSERT INTO booking_services
                                       (booking_id,service_id,service_name,
                                        duration_minutes,price_cents,currency,
                                        sort_order)
                                       VALUES (?,?,?,?,?,'INR',?)""",
                                    (booking_id, service_map.get(key(name)), name, duration if order == 0 else 0, money_cents(row[columns["Price"]]) if order == 0 else 0, order),
                                )
                            report["appointments_added"] += 1
                            report[f"status_{status}"] += 1
                    except Exception as exc:
                        errors.append({"sheet": sheet.title, "row": row_number, "id": source_id, "error": str(exc)})
                        raise
    finally:
        wb.close()
    return dict(report), errors


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--backup", type=Path, required=True)
    parser.add_argument("--clients", type=Path, required=True)
    parser.add_argument("--branch", default="A2Z Technical")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    for path in (args.backup, args.clients):
        if not path.is_file(): raise SystemExit(f"Missing source file: {path}")
    if not postgres_url(): raise SystemExit("A2Z_POSTGRES_URL is not configured")
    preview = {"backup": workbook_totals(args.backup), "clients": workbook_totals(args.clients)}
    with get_db() as conn:
        branch = conn.execute("SELECT id,name FROM branches WHERE lower(name)=lower(?)", (args.branch,)).fetchone()
        if not branch: raise SystemExit(f"Target branch not found: {args.branch}")
        before = {
            "branch_id": branch["id"],
            "instructors": conn.execute("SELECT count(*) AS n FROM instructors WHERE branch_id=?", (branch["id"],)).fetchone()["n"],
            "clients": conn.execute("SELECT count(*) AS n FROM users WHERE branch_id=? AND role='student'", (branch["id"],)).fetchone()["n"],
            "bookings": conn.execute("SELECT count(*) AS n FROM bookings WHERE branch_id=?", (branch["id"],)).fetchone()["n"],
        }
    if not args.apply:
        print(json.dumps({"mode": "preview", "target": before, "source": preview}, indent=2))
        print("Preview only; PostgreSQL was not changed. Run again with --apply after verifying the pg_dump backup.")
        return 0
    counts, errors = apply_import(args.backup, args.clients, args.branch)
    result = {"mode": "applied", "target_before": before, "source": preview, "counts": counts, "errors": errors}
    if args.report: args.report.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(json.dumps(result, indent=2))
    return 2 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
