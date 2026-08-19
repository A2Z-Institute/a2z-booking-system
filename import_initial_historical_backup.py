"""One-time import of the historical Smart Scheduling backup and activity log.

The import is intentionally idempotent through stable source_reference values.
It runs only when invoked by docker-entrypoint on a database with no bookings.
"""

from __future__ import annotations
import json, os, re, sqlite3, sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

DB = Path(os.environ.get("A2Z_DATABASE", "/data/a2z_booking.db"))
BACKUP = Path(os.environ.get("A2Z_INITIAL_BOOKING_BACKUP", "/app/data-import/booking-backup.xlsx"))
ACTIVITY = Path(os.environ.get("A2Z_INITIAL_ACTIVITY_LOG", "/app/data-import/activity-log.xlsx"))
MARKER = Path(os.environ.get("A2Z_INITIAL_IMPORT_MARKER", str(DB.parent / ".initial-backup-imported")))
BREAK_NAMES = {"BREAK","BREAKFAST","TEA BREAK","LUNCH BREAK","NO BOOKING","LEAVE","TEA BREAK"}
STATUS_MAP = {
    "": "Approved", "NONE": "Approved", "CONFIRMED": "Approved",
    "NOT CONFIRMED": "Pending", "PENDING": "Pending",
    "COMPLETED": "Completed", "CANCELLED": "Cancelled",
    "NO-SHOW": "No-show", "NO SHOW": "No-show",
    "RESCHEDULED": "Rescheduled", "RUNNING LATE": "Running Late",
    "ARRIVED": "Arrived",
}
COLORS = ("#E6C8F5","#7CC5E3","#F5E39B","#B9E5BF","#F1DCC0","#D4EC8A")

def clean(v):
    if v is None: return ""
    return " ".join(str(v).replace("\u202f"," ").replace("\xa0"," ").replace("_x000D_"," / ").split()).strip()

def key(v):
    value=clean(v).upper()
    value=value.replace("2O TRAILER","20 TRAILER").replace("COUPLEING","COUPLING").replace("TRAILOR","TRAILER")
    return re.sub(r"[^A-Z0-9]+"," ",value).strip()

def parse_dt(v):
    if isinstance(v, datetime): return v
    if isinstance(v, date): return datetime.combine(v, datetime.min.time())
    raw=clean(v)
    for fmt in ("%d/%m/%Y %I:%M %p","%d/%m/%Y %H:%M","%Y-%m-%d %H:%M:%S"):
        try: return datetime.strptime(raw,fmt)
        except ValueError: pass
    # Handle occasional pandas/openpyxl strings with a single-space AM/PM already normalized.
    raise ValueError(f"Unrecognised date/time {v!r}")

def money_cents(v):
    raw=re.sub(r"[^0-9.-]","",clean(v))
    try: return int(round(float(raw or 0)*100))
    except ValueError: return 0

def split_names(v):
    return [clean(p) for p in re.split(r"\s*[;/]\s*|\s*\n\s*",clean(v)) if clean(p)]

def main():
    if not BACKUP.exists(): raise SystemExit(f"Missing backup workbook: {BACKUP}")
    if not ACTIVITY.exists(): raise SystemExit(f"Missing activity workbook: {ACTIVITY}")
    DB.parent.mkdir(parents=True,exist_ok=True)

    conn=sqlite3.connect(DB, timeout=120)
    conn.row_factory=sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=120000")
    counts=Counter()
    errors=[]

    # Branch/resources already exist after app database initialization.
    branch=conn.execute("SELECT id FROM branches WHERE is_active=1 ORDER BY id LIMIT 1").fetchone()
    if not branch: raise SystemExit("No active branch exists.")
    branch_id=branch["id"]
    admin=conn.execute("SELECT id FROM users WHERE role='admin' ORDER BY id LIMIT 1").fetchone()
    admin_id=admin["id"] if admin else None

    wb=load_workbook(BACKUP, read_only=True, data_only=True)

    try:
        conn.execute("BEGIN IMMEDIATE")

        # 1) Instructors: one real calendar staff record for every backup sheet/provider.
        instructor_names=[]
        for ws in wb.worksheets:
            if ws.title not in {"Clients","Services"}:
                instructor_names.append(ws.title)
        # Providers in Services may include names that are not sheets; include them too.
        if "Services" in wb.sheetnames:
            sws=wb["Services"]
            srows=sws.iter_rows(values_only=True)
            headers=[clean(v) for v in next(srows)]
            six={h:i for i,h in enumerate(headers)}
            for row in srows:
                for provider in split_names(row[six.get("Providers",6)] if "Providers" in six else ""):
                    if provider and provider not in instructor_names:
                        instructor_names.append(provider)

        instructors={}
        for name in instructor_names:
            conn.execute("""INSERT OR IGNORE INTO instructors
                (name,branch_id,specialty,is_active,verification_status)
                VALUES (?,?,NULL,1,'verified')""",(name,branch_id))
            row=conn.execute("SELECT id FROM instructors WHERE branch_id=? AND name=?",(branch_id,name)).fetchone()
            instructors[key(name)]=row["id"]
        counts["instructors"]=len(instructors)

        # 2) Clients.
        if "Clients" in wb.sheetnames:
            ws=wb["Clients"]; rows=ws.iter_rows(values_only=True)
            headers=[clean(v) for v in next(rows)]; ix={h:i for i,h in enumerate(headers)}
            for rn,row in enumerate(rows,2):
                source_id=clean(row[ix.get("Id")]) or f"row-{rn}"
                stable=re.sub(r"[^0-9A-Za-z_-]+","-",source_id).strip("-") or str(rn)
                username=f"smart-client-{stable}".lower()
                first=clean(row[ix.get("First Name")]); last=clean(row[ix.get("Last Name")])
                full_name=" ".join(x for x in (first,last) if x) or f"Client {source_id}"
                email=clean(row[ix.get("Email")]) or None
                phone=clean(row[ix.get("Phone")]) or None
                secondary_email=clean(row[ix.get("Secondary Email")]) or None
                secondary_phone=clean(row[ix.get("Secondary Phone")]) or None
                gender=clean(row[ix.get("Gender")]).lower()
                gender=gender if gender in {"male","female","other"} else None
                birthday=clean(row[ix.get("Birthday")]) or None
                created=clean(row[ix.get("Created")]) or None
                existing=conn.execute("SELECT id FROM users WHERE lower(username)=lower(?)",(username,)).fetchone()
                if existing:
                    uid=existing["id"]
                    conn.execute("""UPDATE users SET full_name=?,email=?,phone=?,branch_id=?,
                        is_active=1,login_enabled=0,must_change_password=0,updated_at=CURRENT_TIMESTAMP
                        WHERE id=?""",(full_name,email,phone,branch_id,uid))
                    counts["clients_updated"]+=1
                else:
                    uid=conn.execute("""INSERT INTO users
                        (username,password_hash,role,full_name,email,phone,branch_id,is_active,login_enabled,must_change_password,created_at,updated_at)
                        VALUES (?,'!','student',?,?,?,?,1,0,0,COALESCE(?,CURRENT_TIMESTAMP),CURRENT_TIMESTAMP)""",
                        (username,full_name,email,phone,branch_id,created)).lastrowid
                    counts["clients_created"]+=1
                conn.execute("""INSERT INTO client_profiles
                    (user_id,secondary_phone,secondary_email,birthday,gender,zip_code,city,street,internal_notes,
                     reminders_enabled,preferred_channel,source_reference,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,1,?,?,CURRENT_TIMESTAMP)
                    ON CONFLICT(user_id) DO UPDATE SET
                    secondary_phone=excluded.secondary_phone,secondary_email=excluded.secondary_email,
                    birthday=excluded.birthday,gender=excluded.gender,zip_code=excluded.zip_code,
                    city=excluded.city,street=excluded.street,internal_notes=excluded.internal_notes,
                    preferred_channel=excluded.preferred_channel,source_reference=excluded.source_reference,
                    updated_at=CURRENT_TIMESTAMP""",
                    (uid,secondary_phone,secondary_email,birthday,gender,
                     clean(row[ix.get("Zip Code")]) or None,clean(row[ix.get("City")]) or None,
                     clean(row[ix.get("Street")]) or None,clean(row[ix.get("Notes")]) or None,
                     "email" if email else "sms",source_id))

        # Client lookup for appointment rows.
        by_phone={}; by_email={}; by_name=defaultdict(list)
        for r in conn.execute("SELECT id,full_name,phone,email FROM users WHERE role='student'"):
            if key(r["phone"]): by_phone[key(r["phone"])]=r["id"]
            if key(r["email"]): by_email[key(r["email"])]=r["id"]
            by_name[key(r["full_name"])].append(r["id"])

        def client_id_for(name,phone,email,source_ref):
            uid=by_phone.get(key(phone)) or by_email.get(key(email))
            if not uid and len(by_name.get(key(name),[]))==1: uid=by_name[key(name)][0]
            if uid: return uid
            username="smart-appt-"+re.sub(r"[^a-z0-9]+","-",source_ref.lower()).strip("-")
            conn.execute("""INSERT OR IGNORE INTO users
                (username,password_hash,role,full_name,email,phone,branch_id,is_active,login_enabled,must_change_password)
                VALUES (?,'!','student',?,?,?,?,1,0,0)""",
                (username,name or "Imported client",email or None,phone or None,branch_id))
            uid=conn.execute("SELECT id FROM users WHERE username=?",(username,)).fetchone()["id"]
            conn.execute("""INSERT OR IGNORE INTO client_profiles
                (user_id,reminders_enabled,preferred_channel,source_reference)
                VALUES (?,0,?,?)""",(uid,"email" if email else "sms",source_ref))
            by_name[key(name)].append(uid)
            if key(phone): by_phone[key(phone)]=uid
            if key(email): by_email[key(email)]=uid
            counts["appointment_clients_created"]+=1
            return uid

        # 3) Services and service/provider/equipment relationships.
        service_map={}
        if "Services" in wb.sheetnames:
            ws=wb["Services"]; rows=ws.iter_rows(values_only=True)
            headers=[clean(v) for v in next(rows)]; ix={h:i for i,h in enumerate(headers)}
            machines=list(conn.execute("SELECT * FROM machines WHERE branch_id=? AND is_active=1",(branch_id,)))
            machine_by_key={key(r["machine_code"]):r for r in machines}
            def machine_for(text):
                target=key(text)
                if target in machine_by_key: return machine_by_key[target]
                candidates=[r for mk,r in machine_by_key.items() if mk in target or target in mk]
                if candidates: return sorted(candidates,key=lambda r:len(key(r["machine_code"])),reverse=True)[0]
                return machines[0] if machines else None

            for pos,row in enumerate(rows):
                name=clean(row[ix["Service"]])
                if not name: continue
                category=clean(row[ix["Group"]]) or "Imported"
                duration_raw=clean(row[ix["Duration"]]); match=re.search(r"\d+",duration_raw)
                duration=int(match.group()) if match else 30
                unavailable={key(x) for x in re.split(r"[,;/]",clean(row[ix["Unavailable"]])) if clean(x)}
                weekdays=",".join(str(n) for n,day in enumerate(("MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY","SUNDAY")) if day not in unavailable)
                conn.execute("""INSERT OR IGNORE INTO services
                    (branch_id,name,description,category,duration_minutes,price_cents,currency,buffer_before_minutes,buffer_after_minutes,color,available_weekdays,requires_approval,is_active)
                    VALUES (?,?,?,?,?,?,'INR',0,0,?,?,0,1)""",
                    (branch_id,name,"Imported from historical booking backup",category,duration,money_cents(row[ix["Price"]]),COLORS[pos%len(COLORS)],weekdays or ""))
                conn.execute("""UPDATE services SET category=?,duration_minutes=?,price_cents=?,available_weekdays=?,is_active=1,updated_at=CURRENT_TIMESTAMP
                    WHERE branch_id=? AND name=?""",(category,duration,money_cents(row[ix["Price"]]),weekdays or "",branch_id,name))
                sid=conn.execute("SELECT id FROM services WHERE branch_id=? AND name=?",(branch_id,name)).fetchone()["id"]
                service_map[key(name)]=sid
                machine=machine_for(name)
                if machine and key(name) not in BREAK_NAMES:
                    conn.execute("INSERT OR IGNORE INTO service_machines(service_id,machine_id) VALUES (?,?)",(sid,machine["id"]))
                for provider in split_names(row[ix["Providers"]]):
                    iid=instructors.get(key(provider))
                    if iid: conn.execute("INSERT OR IGNORE INTO service_instructors(service_id,instructor_id) VALUES (?,?)",(sid,iid))
                counts["services"]+=1

        # 4) Appointments, booking slots, and breaks.
        machines=list(conn.execute("SELECT * FROM machines WHERE branch_id=? AND is_active=1",(branch_id,)))
        machine_by_key={key(r["machine_code"]):r for r in machines}
        def machine_for(text):
            target=key(text)
            if target in machine_by_key: return machine_by_key[target]
            candidates=[r for mk,r in machine_by_key.items() if mk in target or target in mk]
            if candidates: return sorted(candidates,key=lambda r:len(key(r["machine_code"])),reverse=True)[0]
            return machines[0] if machines else None

        for ws in wb.worksheets:
            if ws.title in {"Clients","Services"}: continue
            sheet_key=key(ws.title)
            iid=instructors.get(sheet_key)
            if not iid:
                matches=[v for k,v in instructors.items() if k.startswith(sheet_key) or sheet_key.startswith(k[:20])]
                iid=matches[0] if len(set(matches))==1 else None
            if not iid:
                errors.append({"sheet":ws.title,"error":"No matching instructor"}); continue

            rows=ws.iter_rows(values_only=True)
            headers=[clean(v) for v in next(rows)]; ix={h:i for i,h in enumerate(headers)}
            for rn,row in enumerate(rows,2):
                if not any(v not in (None,"") for v in row): continue
                source_id=clean(row[ix["Id"]]) or f"row-{rn}"
                source_ref=f"smart:{ws.title}:{source_id}"
                try:
                    start=parse_dt(row[ix["Start"]]); end=parse_dt(row[ix["End"]])
                    if end<=start: raise ValueError("end is not after start")
                    client=clean(row[ix["Client"]]); phone=clean(row[ix["Phone"]]); email=clean(row[ix["Email"]])
                    service_text=clean(row[ix["Task/Services"]])
                    status_raw=key(row[ix["Status"]]); notes=clean(row[ix["Notes"]]); custom=clean(row[ix["Custom Fields"]])
                    if custom: notes=(notes+"\n" if notes else "")+"Custom fields: "+custom
                    duration=int((end-start).total_seconds()//60)
                    names=split_names(service_text)
                    primary_service=names[0] if names else "Unspecified training"
                    client_key=key(client)
                    machine=machine_for(primary_service if primary_service!="Unspecified training" else client)
                    if not machine: raise ValueError("No equipment exists")
                    is_slot=("SLOT" in client_key) or (not client and duration>=180 and key(primary_service) not in BREAK_NAMES)
                    is_busy=key(primary_service) in BREAK_NAMES or client_key in BREAK_NAMES
                    sid=service_map.get(key(primary_service))
                    if is_slot:
                        conn.execute("""INSERT INTO booking_slots
                            (instructor_id,machine_id,branch_id,target_date,start_time,end_time,notes,series_id,repeat_rule,created_by,source_reference)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                            ON CONFLICT(source_reference) DO UPDATE SET instructor_id=excluded.instructor_id,machine_id=excluded.machine_id,
                            target_date=excluded.target_date,start_time=excluded.start_time,end_time=excluded.end_time,notes=excluded.notes,updated_at=CURRENT_TIMESTAMP""",
                            (iid,machine["id"],branch_id,start.date().isoformat(),start.strftime("%H:%M"),end.strftime("%H:%M"),
                             notes or client,clean(row[ix["Series-Id"]]) or None,None,admin_id,source_ref))
                        counts["booking_slots"]+=1
                    elif is_busy:
                        conn.execute("""INSERT INTO instructor_time_off
                            (instructor_id,target_date,start_time,end_time,reason,notes,series_id,created_by,source_reference)
                            VALUES (?,?,?,?,?,?,?,?,?)
                            ON CONFLICT(source_reference) DO UPDATE SET instructor_id=excluded.instructor_id,target_date=excluded.target_date,
                            start_time=excluded.start_time,end_time=excluded.end_time,reason=excluded.reason,notes=excluded.notes,updated_at=CURRENT_TIMESTAMP""",
                            (iid,start.date().isoformat(),start.strftime("%H:%M"),end.strftime("%H:%M"),
                             primary_service or client,notes,clean(row[ix["Series-Id"]]) or None,admin_id,source_ref))
                        counts["breaks_busy"]+=1
                    else:
                        uid=client_id_for(client,phone,email,source_ref)
                        mapped=STATUS_MAP.get(status_raw,"Pending")
                        cur=conn.execute("""INSERT INTO bookings
                            (student_name,mobile_number,machine_id,instructor_id,branch_id,target_date,start_time,end_time,
                             validation_status,student_user_id,notes,service_id,service_name,service_price_cents,currency,
                             buffer_before_minutes,buffer_after_minutes,series_id,allow_double_booking,source_reference)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'INR',0,0,?,1,?)
                            ON CONFLICT(source_reference) DO UPDATE SET
                             student_name=excluded.student_name,mobile_number=excluded.mobile_number,machine_id=excluded.machine_id,
                             instructor_id=excluded.instructor_id,branch_id=excluded.branch_id,target_date=excluded.target_date,
                             start_time=excluded.start_time,end_time=excluded.end_time,validation_status=excluded.validation_status,
                             student_user_id=excluded.student_user_id,notes=excluded.notes,service_id=excluded.service_id,
                             service_name=excluded.service_name,service_price_cents=excluded.service_price_cents,
                             series_id=excluded.series_id,updated_at=CURRENT_TIMESTAMP
                            RETURNING id""",
                            (client or "Imported appointment",phone,machine["id"],iid,branch_id,start.date().isoformat(),
                             start.strftime("%H:%M"),end.strftime("%H:%M"),mapped,uid,notes,sid,service_text,
                             money_cents(row[ix["Price"]]),clean(row[ix["Series-Id"]]) or None,source_ref))
                        bid=cur.fetchone()["id"]
                        conn.execute("DELETE FROM booking_services WHERE booking_id=?",(bid,))
                        for order,name in enumerate(names or [primary_service]):
                            service_id=service_map.get(key(name))
                            conn.execute("""INSERT INTO booking_services
                                (booking_id,service_id,service_name,duration_minutes,price_cents,currency,sort_order)
                                VALUES (?,?,?,?,?,'INR',?)""",
                                (bid,service_id,name,duration if order==0 else 0,
                                 money_cents(row[ix["Price"]]) if order==0 else 0,order))
                        counts["appointments"]+=1
                        counts[f"status_{mapped}"]+=1
                except Exception as exc:
                    errors.append({"sheet":ws.title,"row":rn,"id":source_id,"error":str(exc)})
                    counts["errors"]+=1

        # 5) Historical activity log. Preserve the original actor even though
        # these accounts are disabled and cannot log in.
        activity_wb=load_workbook(ACTIVITY,read_only=True,data_only=True)
        aws=activity_wb.active
        rows=aws.iter_rows(values_only=True)
        headers=[clean(v) for v in next(rows)]; ix={h:i for i,h in enumerate(headers)}
        actor_ids={}
        for row in rows:
            username=clean(row[ix["User"]]) or "system-import"
            if username not in actor_ids:
                existing=conn.execute("SELECT id FROM users WHERE lower(username)=lower(?)",(username,)).fetchone()
                if existing:
                    actor_ids[username]=existing["id"]
                else:
                    actor_ids[username]=conn.execute("""INSERT INTO users
                        (username,password_hash,role,full_name,branch_id,is_active,login_enabled,must_change_password)
                        VALUES (?,'!','booking_agent',?,?,0,0,0)""",
                        (username,username,branch_id)).lastrowid

        # Re-read rows because the iterator was consumed.
        rows=aws.iter_rows(values_only=True); next(rows)
        # Map legacy numeric booking IDs to current booking IDs when possible.
        booking_map={}
        for r in conn.execute("SELECT id,source_reference FROM bookings WHERE source_reference IS NOT NULL"):
            tail=r["source_reference"].rsplit(":",1)[-1]
            if tail.isdigit(): booking_map[tail]=r["id"]

        for row in rows:
            try:
                raw_dt=clean(row[ix["Date/Time"]])
                created_at=parse_dt(raw_dt).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username=clean(row[ix["User"]]) or "system-import"
            action=clean(row[ix["Action"]]) or "UNKNOWN"
            typ=clean(row[ix["Type"]]) or "Legacy"
            details_id=clean(row[ix["ID / Details"]])
            ip=clean(row[ix["IP Address"]])
            booking_id=booking_map.get(details_id) if details_id.isdigit() else None
            details={"legacy_action":action,"legacy_type":typ,"legacy_id_or_details":details_id,"legacy_ip_address":ip}
            # Stable source reference is encoded in details so duplicate runs can be detected
            # by the marker; the import is one-time by design.
            conn.execute("""INSERT INTO audit_events
                (actor_user_id,booking_id,event_type,details,created_at)
                VALUES (?,?,?,?,?)""",
                (actor_ids.get(username),booking_id,f"legacy_{action.lower()}_{typ.lower()}",
                 json.dumps(details,ensure_ascii=True),created_at))
            counts["activity_events"]+=1
        activity_wb.close()

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        wb.close()
        conn.close()

    # Marker is created only after a successful transaction.
    MARKER.write_text(json.dumps({"completed_at":datetime.now().isoformat(timespec="seconds"),"counts":dict(counts)},indent=2),encoding="utf-8")
    print(json.dumps({"success":True,"counts":dict(counts),"errors":errors},indent=2))
    if errors:
        raise SystemExit(2)

if __name__=="__main__":
    main()
